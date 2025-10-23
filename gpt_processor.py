#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OpenAI GPTæ‰¹å¤„ç†æœ¯è¯­æŠ½å–å¤„ç†å™¨
æ”¯æŒå¹¶å‘æ‰¹é‡æ–‡æœ¬å¤„ç†ã€ä»»åŠ¡ç›‘æ§å’Œç»“æœå¤„ç†
"""

import json
import time
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from checkpoint_manager import CheckpointManager

try:
    from openai import OpenAI
    import tiktoken
except ImportError:
    print("è¯·å®‰è£…OpenAIåº“: pip install openai tiktoken")
    raise


class GPTProcessor:
    """OpenAI GPTæ‰¹å¤„ç†å¤„ç†å™¨"""
    
    def __init__(self, 
                 api_key: str, 
                 base_url: str = "https://api.openai.com/v1", 
                 base_dir: str = "batch_results",
                 enable_checkpoint: bool = True,
                 checkpoint_dir: str = "checkpoints"):
        """
        åˆå§‹åŒ–æ‰¹å¤„ç†å™¨
        
        Args:
            api_key: OpenAI APIå¯†é’¥
            base_url: APIç«¯ç‚¹URL
            base_dir: ç»“æœå­˜å‚¨ç›®å½•
            enable_checkpoint: æ˜¯å¦å¯ç”¨æ–­ç‚¹åŠŸèƒ½
            checkpoint_dir: æ–­ç‚¹æ–‡ä»¶ç›®å½•
        """
        self.client = OpenAI(
            api_key=api_key, 
            base_url=base_url,
            timeout=60.0,  # è®¾ç½®60ç§’è¶…æ—¶
            max_retries=3   # æœ€å¤§é‡è¯•3æ¬¡
        )
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(exist_ok=True)
        
        # åˆå§‹åŒ–æ–­ç‚¹ç®¡ç†å™¨
        self.enable_checkpoint = enable_checkpoint
        self.checkpoint_manager = CheckpointManager(checkpoint_dir) if enable_checkpoint else None
        
        # é…ç½®æ—¥å¿—å’Œå¹¶å‘æ§åˆ¶
        self._setup_logging()
        self._setup_concurrency_control()
    
    # =============================================================================
    # åˆå§‹åŒ–å’Œé…ç½®
    # =============================================================================
    
    def _setup_logging(self):
        """è®¾ç½®æ—¥å¿—é…ç½®"""
        log_file = self.base_dir / "gpt_processor.log"
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def _setup_concurrency_control(self):
        """è®¾ç½®å¹¶å‘æ§åˆ¶"""
        # é™ä½å¹¶å‘æ•°ä»¥å‡å°‘è¿æ¥é—®é¢˜
        self.semaphore = threading.Semaphore(5)  # æœ€å¤§5ä¸ªå¹¶å‘è¯·æ±‚
        self.logger.info("GPTå¤„ç†å™¨åˆå§‹åŒ–å®Œæˆ")
    
    # =============================================================================
    # Tokenè®¡ç®—
    # =============================================================================
    
    def count_tokens(self, text: str, model: str = "gpt-4") -> int:
        """è®¡ç®—æ–‡æœ¬çš„tokenæ•°é‡"""
        try:
            # å°è¯•è·å–æ¨¡å‹ç‰¹å®šçš„ç¼–ç å™¨
            encoding = tiktoken.encoding_for_model(model)
            return len(encoding.encode(text))
        except Exception:
            # å¦‚æœæ¨¡å‹ä¸æ”¯æŒï¼Œä½¿ç”¨é€šç”¨ç¼–ç å™¨
            encoding = tiktoken.get_encoding("cl100k_base")
            return len(encoding.encode(text))
    
    # =============================================================================
    # å•æ–‡æœ¬å¤„ç†
    # =============================================================================
    
    def process_single_text(self, 
                          text: str,
                          custom_id: str,
                          system_prompt: str = None,
                          user_prompt_template: str = None,
                          model: str = "gpt-4-turbo-preview",
                          temperature: float = 0.1,
                          max_tokens: int = 4096,
                          source_file: str = None) -> Dict[str, Any]:
        """
        å¤„ç†å•ä¸ªæ–‡æœ¬
        
        Args:
            text: è¦å¤„ç†çš„æ–‡æœ¬
            custom_id: è‡ªå®šä¹‰ID
            system_prompt: ç³»ç»Ÿæç¤ºè¯
            user_prompt_template: ç”¨æˆ·æç¤ºè¯æ¨¡æ¿
            model: æ¨¡å‹åç§°
            temperature: æ¸©åº¦å‚æ•°
            max_tokens: æœ€å¤§è¾“å‡ºtokenæ•°
            source_file: æ¥æºæ–‡ä»¶å
            
        Returns:
            å¤„ç†ç»“æœå­—å…¸
        """
        with self.semaphore:  # æ§åˆ¶å¹¶å‘æ•°
            try:
                # éªŒè¯å¿…è¦å‚æ•°
                if system_prompt is None:
                    raise ValueError("system_prompt is required")
                
                if user_prompt_template is None:
                    raise ValueError("user_prompt_template is required")
                
                # æ„å»ºAPIè°ƒç”¨å‚æ•°
                api_params = self._build_api_params(
                    system_prompt, user_prompt_template, text, model, temperature, max_tokens
                )
                
                # è®°å½•å¤„ç†ä¿¡æ¯
                total_tokens = self.count_tokens(system_prompt + user_prompt_template.format(text=text), model)
                self.logger.info(f"å¤„ç† {custom_id}: è¾“å…¥ {total_tokens} tokens")
                
                # è°ƒç”¨API
                response = self.client.chat.completions.create(**api_params)
                
                # å¤„ç†å“åº”
                return self._process_api_response(response, custom_id, model, source_file)
                
            except Exception as e:
                self.logger.error(f"âŒ {custom_id} å¤„ç†å¤±è´¥: {e}")
                return self._create_error_result(custom_id, model, source_file, str(e))
    
    def _build_api_params(self, system_prompt: str, user_prompt_template: str, text: str, 
                         model: str, temperature: float, max_tokens: int) -> Dict[str, Any]:
        """æ„å»ºAPIè°ƒç”¨å‚æ•°"""
        from config import get_token_param_name
        
        user_prompt = user_prompt_template.format(text=text)
        
        api_params = {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": temperature
        }
        
        # æ ¹æ®æ¨¡å‹ç±»å‹ä½¿ç”¨æ­£ç¡®çš„tokenå‚æ•°
        token_param = get_token_param_name(model)
        api_params[token_param] = max_tokens
        
        return api_params
    
    def _process_api_response(self, response, custom_id: str, model: str, source_file: str) -> Dict[str, Any]:
        """å¤„ç†APIå“åº”"""
        # å¤„ç†ä¸åŒç±»å‹çš„å“åº”
        if isinstance(response, str):
            content = response
            usage_info = {"total_tokens": 0}
            model_name = model
        else:
            content = response.choices[0].message.content
            usage_info = {
                "total_tokens": response.usage.total_tokens if response.usage else 0,
                "prompt_tokens": response.usage.prompt_tokens if response.usage else 0,
                "completion_tokens": response.usage.completion_tokens if response.usage else 0
            }
            model_name = response.model
        
        # è§£æJSONå“åº”
        extracted_terms = self._parse_json_response(content)
        
        result = {
            "custom_id": custom_id,
            "extracted_terms": extracted_terms,
            "usage": usage_info,
            "model": model_name,
            "source_file": source_file,
            "created": int(time.time())
        }
        
        self.logger.info(f"âœ… {custom_id} å¤„ç†å®Œæˆ: {usage_info.get('total_tokens', 0)} tokens")
        return result
    
    def _parse_json_response(self, content: str) -> Dict[str, Any]:
        """è§£æJSONå“åº”å†…å®¹"""
        try:
            # å°è¯•æŸ¥æ‰¾JSONå—
            json_start = content.find('{')
            if json_start == -1:
                json_start = content.find('[')
            
            if json_start != -1:
                json_end = content.rfind('}')
                if json_end == -1:
                    json_end = content.rfind(']')
                if json_end != -1:
                    content = content[json_start:json_end+1].strip()
                else:
                    # å¦‚æœæ‰¾ä¸åˆ°ç»“æŸç¬¦ï¼Œä»å¼€å§‹ç¬¦åˆ°æœ€å
                    json_end = content.rfind('}')
                    if json_end != -1:
                        content = content[json_start:json_end].strip()
            
            return json.loads(content)
        except json.JSONDecodeError:
            return {"raw_content": content}
    
    def _create_error_result(self, custom_id: str, model: str, source_file: str, error_msg: str) -> Dict[str, Any]:
        """åˆ›å»ºé”™è¯¯ç»“æœ"""
        return {
            "custom_id": custom_id,
            "error": error_msg,
            "extracted_terms": {"raw_content": f"å¤„ç†å¤±è´¥: {error_msg}"},
            "usage": {"total_tokens": 0},
            "model": model,
            "source_file": source_file,
            "created": int(time.time())
        }
    
    # =============================================================================
    # æ‰¹é‡å¹¶å‘å¤„ç†
    # =============================================================================
    
    def process_batch_concurrent(self, 
                               texts: List[str],
                               system_prompt: str = None,
                               user_prompt_template: str = None,
                               model: str = "gpt-4-turbo-preview",
                               temperature: float = 0.1,
                               max_tokens: int = 4096,
                               max_concurrent: int = 10,
                               source_files: List[str] = None) -> List[Dict[str, Any]]:
        """
        å¹¶å‘å¤„ç†æ‰¹é‡æ–‡æœ¬
        
        Args:
            texts: è¦å¤„ç†çš„æ–‡æœ¬åˆ—è¡¨
            system_prompt: ç³»ç»Ÿæç¤ºè¯
            user_prompt_template: ç”¨æˆ·æç¤ºè¯æ¨¡æ¿
            model: æ¨¡å‹åç§°
            temperature: æ¸©åº¦å‚æ•°
            max_tokens: æœ€å¤§è¾“å‡ºtokenæ•°
            max_concurrent: æœ€å¤§å¹¶å‘æ•°
            source_files: æ¥æºæ–‡ä»¶ååˆ—è¡¨
            
        Returns:
            å¤„ç†ç»“æœåˆ—è¡¨
        """
        self.logger.info(f"ğŸš€ å¼€å§‹å¹¶å‘æ‰¹å¤„ç†: {len(texts)} ä¸ªæ–‡æœ¬ï¼Œæœ€å¤§å¹¶å‘ {max_concurrent}")
        
        # æ›´æ–°å¹¶å‘æ§åˆ¶
        self.semaphore = threading.Semaphore(max_concurrent)
        
        results = []
        
        with ThreadPoolExecutor(max_workers=max_concurrent) as executor:
            # æäº¤æ‰€æœ‰ä»»åŠ¡
            future_to_id = self._submit_batch_tasks(
                executor, texts, system_prompt, user_prompt_template, 
                model, temperature, max_tokens, source_files
            )
            
            # æ”¶é›†ç»“æœ
            results = self._collect_batch_results(future_to_id, len(texts), model, source_files)
        
        # æŒ‰custom_idæ’åºç»“æœ
        results.sort(key=lambda x: x.get("custom_id", ""))
        
        self.logger.info(f"âœ… å¹¶å‘æ‰¹å¤„ç†å®Œæˆ: {len(results)} ä¸ªç»“æœ")
        return results
    
    def _submit_batch_tasks(self, executor, texts: List[str], system_prompt: str, 
                           user_prompt_template: str, model: str, temperature: float, 
                           max_tokens: int, source_files: List[str]) -> Dict:
        """æäº¤æ‰¹å¤„ç†ä»»åŠ¡"""
        future_to_id = {}
        
        for i, text in enumerate(texts):
            custom_id = f"term-extraction-{i+1}"
            source_file = source_files[i] if source_files and i < len(source_files) else None
            
            future = executor.submit(
                self.process_single_text,
                text=text,
                custom_id=custom_id,
                system_prompt=system_prompt,
                user_prompt_template=user_prompt_template,
                model=model,
                temperature=temperature,
                max_tokens=max_tokens,
                source_file=source_file
            )
            future_to_id[future] = custom_id
        
        return future_to_id
    
    def _collect_batch_results(self, future_to_id: Dict, total_count: int, 
                              model: str, source_files: List[str]) -> List[Dict[str, Any]]:
        """æ”¶é›†æ‰¹å¤„ç†ç»“æœ"""
        results = []
        completed_count = 0
        
        for future in as_completed(future_to_id):
            custom_id = future_to_id[future]
            try:
                result = future.result()
                results.append(result)
                completed_count += 1
                self.logger.info(f"ğŸ“Š è¿›åº¦: {completed_count}/{total_count} å®Œæˆ")
                
                # æ›´æ–°æ–­ç‚¹çŠ¶æ€ - æˆåŠŸ
                text_index = int(custom_id.split('-')[-1]) - 1
                source_file = source_files[text_index] if source_files and text_index < len(source_files) else f"text_{text_index+1}.txt"
                self.update_text_processing_status(
                    text_index=text_index,
                    source_file=source_file,
                    status='completed',
                    result_data=result
                )
                
            except Exception as e:
                self.logger.error(f"âŒ {custom_id} å¤„ç†å¼‚å¸¸: {e}")
                # æ·»åŠ é”™è¯¯ç»“æœ
                text_index = int(custom_id.split('-')[-1]) - 1
                source_file = source_files[text_index] if source_files and text_index < len(source_files) else f"text_{text_index+1}.txt"
                results.append(self._create_error_result(custom_id, model, source_file, str(e)))
                
                # æ›´æ–°æ–­ç‚¹çŠ¶æ€ - å¤±è´¥
                self.update_text_processing_status(
                    text_index=text_index,
                    source_file=source_file,
                    status='failed',
                    error_message=str(e)
                )
        
        return results
    
    # =============================================================================
    # æœ¯è¯­å»é‡å’Œåˆå¹¶
    # =============================================================================
    
    def deduplicate_terms(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        å»é‡å’Œåˆå¹¶ç›¸åŒæœ¯è¯­
        
        Args:
            results: å¤„ç†ç»“æœåˆ—è¡¨
            
        Returns:
            å»é‡åçš„ç»“æœåˆ—è¡¨
        """
        from config import TERM_PROCESSING
        
        self.logger.info("å¼€å§‹æœ¯è¯­å»é‡å¤„ç†...")
        
        # æ”¶é›†æ‰€æœ‰æœ¯è¯­
        all_terms = self._collect_all_terms(results)
        
        # åˆå¹¶é‡å¤æœ¯è¯­
        merged_terms, duplicate_count = self._merge_duplicate_terms(all_terms)
        
        # åˆ›å»ºåˆå¹¶ç»“æœ
        merged_result = self._create_merged_result(results, merged_terms, duplicate_count)
        
        self.logger.info(f"å»é‡å®Œæˆ: {len(merged_terms)} ä¸ªå”¯ä¸€æœ¯è¯­ (ç§»é™¤ {duplicate_count} ä¸ªé‡å¤)")
        return [merged_result]
    
    def _collect_all_terms(self, results: List[Dict[str, Any]]) -> Dict[str, List[Dict]]:
        """æ”¶é›†æ‰€æœ‰æœ¯è¯­ï¼ˆæ”¯æŒåŒè¯­ï¼‰"""
        from config import TERM_PROCESSING
        
        all_terms = {}  # {term_key: [term_objects]}
        case_sensitive = TERM_PROCESSING.get("case_sensitive_matching", False)
        
        for result in results:
            extracted_terms = result.get("extracted_terms", {})
            if "terms" in extracted_terms and isinstance(extracted_terms["terms"], list):
                for term in extracted_terms["terms"]:
                    if not isinstance(term, dict):
                        continue
                    
                    # æ”¯æŒåŒè¯­æ ¼å¼
                    eng_term = term.get("eng_term", "").strip()
                    zh_term = term.get("zh_term", "").strip()
                    
                    # å…¼å®¹æ—§æ ¼å¼
                    if not eng_term and not zh_term and "term" in term:
                        single_term = term["term"].strip()
                        term_key = single_term if case_sensitive else single_term.lower()
                        
                        if term_key not in all_terms:
                            all_terms[term_key] = []
                        all_terms[term_key].append({
                            "original_term": term["term"],
                            "source_id": result.get("custom_id", ""),
                            "source_file": result.get("source_file", "")
                        })
                    else:
                        # åŒè¯­æ ¼å¼ï¼šä½¿ç”¨è‹±æ–‡+ä¸­æ–‡ç»„åˆä½œä¸ºkey
                        eng_key = eng_term if case_sensitive else eng_term.lower()
                        zh_key = zh_term  # ä¸­æ–‡ä¸éœ€è¦åŒºåˆ†å¤§å°å†™
                        term_key = f"{eng_key}|{zh_key}"
                        
                        if term_key not in all_terms:
                            all_terms[term_key] = []
                        all_terms[term_key].append({
                            "original_eng_term": eng_term,
                            "original_zh_term": zh_term,
                            "source_id": result.get("custom_id", ""),
                            "source_file": result.get("source_file", "")
                        })
        
        return all_terms
    
    def _merge_duplicate_terms(self, all_terms: Dict[str, List[Dict]]) -> Tuple[List[Dict], int]:
        """åˆå¹¶é‡å¤æœ¯è¯­"""
        merged_terms = []
        duplicate_count = 0
        
        for term_name, term_list in all_terms.items():
            if len(term_list) > 1:
                duplicate_count += len(term_list) - 1
            
            # åˆå¹¶æœ¯è¯­ä¿¡æ¯
            merged_term = self._merge_term_info(term_list)
            merged_terms.append(merged_term)
        
        return merged_terms, duplicate_count
    
    def _create_merged_result(self, results: List[Dict[str, Any]], merged_terms: List[Dict], 
                             duplicate_count: int) -> Dict[str, Any]:
        """åˆ›å»ºåˆå¹¶ç»“æœ"""
        return {
            "custom_id": "merged_terms",
            "extracted_terms": {
                "terms": merged_terms,
                "total_terms": len(merged_terms),
                "original_terms": sum(len(term_list) for term_list in self._collect_all_terms(results).values()),
                "duplicates_removed": duplicate_count
            },
            "usage": {"total_tokens": sum(r.get("usage", {}).get("total_tokens", 0) for r in results)},
            "model": "merged",
            "created": max(r.get("created", 0) for r in results) if results else 0
        }
    
    def _merge_term_info(self, term_list: List[Dict]) -> Dict:
        """
        åˆå¹¶ç›¸åŒæœ¯è¯­çš„ä¿¡æ¯ï¼ˆæ”¯æŒåŒè¯­ï¼‰
        
        Args:
            term_list: ç›¸åŒæœ¯è¯­çš„åˆ—è¡¨
            
        Returns:
            åˆå¹¶åçš„æœ¯è¯­ä¿¡æ¯
        """
        # æ”¶é›†æ‰€æœ‰æ¥æºæ–‡ä»¶ï¼ˆå»é‡ï¼‰
        source_files = self._collect_source_files(term_list)
        
        # åˆ¤æ–­æ˜¯åŒè¯­æ ¼å¼è¿˜æ˜¯æ—§æ ¼å¼
        if "original_eng_term" in term_list[0]:
            # åŒè¯­æ ¼å¼
            if len(term_list) == 1:
                return {
                    "eng_term": term_list[0]["original_eng_term"],
                    "zh_term": term_list[0]["original_zh_term"],
                    "source_file": term_list[0].get("source_file", "")
                }
            else:
                # é€‰æ‹©æœ€ä½³çš„æœ¯è¯­ç‰ˆæœ¬
                best_eng_term = self._select_best_bilingual_term(term_list, "original_eng_term")
                best_zh_term = term_list[0]["original_zh_term"]  # ä¸­æ–‡é€šå¸¸ä¸€è‡´
                
                return {
                    "eng_term": best_eng_term,
                    "zh_term": best_zh_term,
                    "source_files": source_files if len(source_files) > 1 else source_files[0] if source_files else ""
                }
        else:
            # æ—§æ ¼å¼ï¼ˆå•ä¸€termå­—æ®µï¼‰
            if len(term_list) == 1:
                return {
                    "term": term_list[0]["original_term"],
                    "source_file": term_list[0].get("source_file", "")
                }
            
            # é€‰æ‹©æœ€ä½³çš„åŸå§‹æœ¯è¯­åï¼ˆä¼˜å…ˆé€‰æ‹©é¦–å­—æ¯å¤§å†™çš„ï¼‰
            best_term = self._select_best_term(term_list)
            
            return {
                "term": best_term,
                "source_files": source_files if len(source_files) > 1 else source_files[0] if source_files else ""
            }
    
    def _select_best_term(self, term_list: List[Dict]) -> str:
        """é€‰æ‹©æœ€ä½³çš„æœ¯è¯­åç§°ï¼ˆæ—§æ ¼å¼ï¼‰"""
        best_term = term_list[0]["original_term"]
        for term_info in term_list:
            if term_info["original_term"] and term_info["original_term"][0].isupper():
                best_term = term_info["original_term"]
                break
        return best_term
    
    def _select_best_bilingual_term(self, term_list: List[Dict], field_name: str) -> str:
        """é€‰æ‹©æœ€ä½³çš„åŒè¯­æœ¯è¯­åç§°ï¼ˆä¼˜å…ˆé€‰æ‹©é¦–å­—æ¯å¤§å†™çš„ï¼‰"""
        best_term = term_list[0][field_name]
        for term_info in term_list:
            term_value = term_info.get(field_name, "")
            if term_value and term_value[0].isupper():
                best_term = term_value
                break
        return best_term
    
    def _collect_source_files(self, term_list: List[Dict]) -> List[str]:
        """æ”¶é›†æ¥æºæ–‡ä»¶åˆ—è¡¨"""
        return list(set([
            term_info.get("source_file", "") 
            for term_info in term_list 
            if term_info.get("source_file")
        ]))
    
    def _extract_source_filename(self, source_files: List[str]) -> str:
        """ä»æºæ–‡ä»¶åˆ—è¡¨ä¸­æå–ä¸»è¦æ–‡ä»¶åï¼ˆä¸å«æ‰©å±•åå’Œè·¯å¾„ï¼‰"""
        if not source_files:
            return "unknown"
        
        # å–ç¬¬ä¸€ä¸ªæ–‡ä»¶ä½œä¸ºä¸»æ–‡ä»¶å
        main_file = source_files[0]
        
        # å¤„ç†è™šæ‹Ÿæ–‡ä»¶åæ ¼å¼ï¼Œå¦‚ "filename.pdf - ç‰‡æ®µ 1/5 (2621 tokens)"
        import re
        # ç§»é™¤ç‰‡æ®µä¿¡æ¯ï¼Œåªä¿ç•™åŸå§‹æ–‡ä»¶åéƒ¨åˆ†
        clean_main_file = re.sub(r'\s*-\s*ç‰‡æ®µ\s*\d+/\d+\s*\([^)]+\)', '', main_file)
        
        # æå–æ–‡ä»¶åéƒ¨åˆ†ï¼Œå»æ‰è·¯å¾„å’Œæ‰©å±•å
        from pathlib import Path
        filename = Path(clean_main_file.strip()).stem
        
        # æ¸…ç†æ–‡ä»¶åä¸­çš„ç‰¹æ®Šå­—ç¬¦ï¼Œåªä¿ç•™ä¸­è‹±æ–‡å­—æ¯ã€æ•°å­—å’Œä¸‹åˆ’çº¿
        clean_filename = re.sub(r'[^\w\u4e00-\u9fff]', '_', filename)
        
        return clean_filename[:50]  # é™åˆ¶é•¿åº¦é¿å…æ–‡ä»¶åè¿‡é•¿
    
    def _count_total_terms(self, results: List[Dict[str, Any]]) -> int:
        """è®¡ç®—åˆå¹¶åçš„æ€»æœ¯è¯­æ•°"""
        total = 0
        for result in results:
            if result.get("custom_id") == "merged_terms":
                extracted_terms = result.get("extracted_terms", {})
                if "terms" in extracted_terms:
                    total = len(extracted_terms["terms"])
                elif "total_terms" in extracted_terms:
                    total = extracted_terms["total_terms"]
                break
        
        return total
    
    # =============================================================================
    # ç»“æœä¿å­˜
    # =============================================================================
    
    def save_processed_results(self, 
                              results: List[Dict[str, Any]], 
                              output_format: str = "json",
                              source_filename: str = "",
                              model_name: str = "",
                              total_terms: int = 0) -> str:
        """
        ä¿å­˜å¤„ç†ç»“æœåˆ°æ–‡ä»¶
        
        Args:
            results: å¤„ç†ç»“æœåˆ—è¡¨
            output_format: è¾“å‡ºæ ¼å¼ (json/csv/txt/excel)
            source_filename: æºæ–‡ä»¶åï¼ˆä¸å«æ‰©å±•åï¼‰
            model_name: ä½¿ç”¨çš„æ¨¡å‹åç§°
            total_terms: åˆå¹¶åçš„æ€»æœ¯è¯­æ•°
            
        Returns:
            è¾“å‡ºæ–‡ä»¶è·¯å¾„
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # æ„å»ºæ–°çš„å‘½åæ ¼å¼: æ–‡ä»¶-æ—¥æœŸ-æ¨¡å‹-æ€»æ¡æ•°
        if source_filename and model_name and total_terms > 0:
            filename_base = f"{source_filename}_{timestamp}_{model_name}_{total_terms}terms"
        else:
            filename_base = f"processed_terms_{timestamp}"
        
        if output_format == "json":
            return self._save_json_results(results, filename_base)
        elif output_format == "csv":
            return self._save_csv_results(results, filename_base)
        elif output_format == "txt":
            return self._save_txt_results(results, filename_base)
        elif output_format == "excel":
            return self._save_excel_results(results, filename_base)
        elif output_format == "tbx":
            return self._save_tbx_results(results, filename_base)
        else:
            raise ValueError(f"ä¸æ”¯æŒçš„è¾“å‡ºæ ¼å¼: {output_format}")
    
    def _save_json_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """ä¿å­˜JSONæ ¼å¼ç»“æœ"""
        output_file = self.base_dir / f"{filename_base}.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"ç»“æœå·²ä¿å­˜åˆ°: {output_file}")
        return str(output_file)
    
    def _save_csv_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """ä¿å­˜CSVæ ¼å¼ç»“æœï¼ˆæ”¯æŒåŒè¯­ï¼‰"""
        import csv
        
        output_file = self.base_dir / f"{filename_base}.csv"
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:  # ä½¿ç”¨utf-8-sigæ”¯æŒExcel
            if results:
                writer = csv.writer(f)
                # å†™å…¥æ ‡é¢˜è¡Œï¼ˆåŒè¯­ï¼‰
                writer.writerow(["custom_id", "è‹±æ–‡æœ¯è¯­", "ä¸­æ–‡æœ¯è¯­", "æ¥æºæ–‡ä»¶", "æ¨¡å‹", "tokenä½¿ç”¨"])
                
                for result in results:
                    terms = result.get("extracted_terms", {})
                    if "terms" in terms and isinstance(terms["terms"], list):
                        for term in terms["terms"]:
                            # æ”¯æŒåŒè¯­æ ¼å¼
                            eng_term = term.get("eng_term", "")
                            zh_term = term.get("zh_term", "")
                            
                            # å…¼å®¹æ—§æ ¼å¼ï¼ˆå•ä¸€termå­—æ®µï¼‰
                            if not eng_term and not zh_term:
                                single_term = term.get("term", "")
                                # ç®€å•åˆ¤æ–­ï¼šå¦‚æœåŒ…å«ä¸­æ–‡å­—ç¬¦ï¼Œæ”¾å…¥ä¸­æ–‡åˆ—
                                if any('\u4e00' <= char <= '\u9fff' for char in single_term):
                                    zh_term = single_term
                                else:
                                    eng_term = single_term
                            
                            writer.writerow([
                                result.get("custom_id", ""),
                                eng_term,
                                zh_term,
                                term.get("source_file", result.get("source_file", "")),
                                result.get("model", ""),
                                result.get("usage", {}).get("total_tokens", 0)
                            ])
                    else:
                        # å…¼å®¹å…¶ä»–æ ¼å¼
                        writer.writerow([
                            result.get("custom_id", ""),
                            "",
                            str(terms.get("raw_content", "")),
                            result.get("source_file", ""),
                            result.get("model", ""),
                            result.get("usage", {}).get("total_tokens", 0)
                        ])
        
        self.logger.info(f"ç»“æœå·²ä¿å­˜åˆ°: {output_file}")
        return str(output_file)
    
    def _save_txt_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """ä¿å­˜TXTæ ¼å¼ç»“æœï¼ˆæ”¯æŒåŒè¯­ï¼‰"""
        output_file = self.base_dir / f"{filename_base}.txt"
        with open(output_file, 'w', encoding='utf-8') as f:
            for i, result in enumerate(results, 1):
                f.write(f"=== ç»“æœ {i}: {result.get('custom_id', 'unknown')} ===\n")
                terms = result.get("extracted_terms", {})
                
                if "terms" in terms and isinstance(terms["terms"], list):
                    f.write(f"æå–çš„æœ¯è¯­æ•°é‡: {len(terms['terms'])}\n")
                    f.write(f"æ¥æºæ–‡ä»¶: {result.get('source_file', 'æœªçŸ¥')}\n")
                    f.write(f"ä½¿ç”¨æ¨¡å‹: {result.get('model', 'æœªçŸ¥')}\n")
                    f.write(f"Tokenä½¿ç”¨: {result.get('usage', {}).get('total_tokens', 0)}\n\n")
                    
                    for j, term in enumerate(terms["terms"], 1):
                        # æ”¯æŒåŒè¯­æ ¼å¼
                        eng_term = term.get("eng_term", "")
                        zh_term = term.get("zh_term", "")
                        
                        # å…¼å®¹æ—§æ ¼å¼
                        if not eng_term and not zh_term:
                            single_term = term.get("term", "")
                            f.write(f"{j}. {single_term}\n")
                        else:
                            # åŒè¯­æ ¼å¼è¾“å‡º
                            f.write(f"{j}. {eng_term}\n")
                            f.write(f"   {zh_term}\n")
                        
                        if term.get('source_file'):
                            f.write(f"   æ¥æº: {term.get('source_file')}\n")
                else:
                    f.write(f"åŸå§‹å†…å®¹: {terms.get('raw_content', 'æ— å†…å®¹')}\n")
                
                f.write("="*50 + "\n\n")
        
        self.logger.info(f"ç»“æœå·²ä¿å­˜åˆ°: {output_file}")
        return str(output_file)
    
    def _save_excel_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """ä¿å­˜Excelæ ¼å¼ç»“æœ"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("éœ€è¦å®‰è£…openpyxlåº“: pip install openpyxl")
        
        output_file = self.base_dir / f"{filename_base}.xlsx"
        
        # åˆ›å»ºå·¥ä½œç°¿å’Œå·¥ä½œè¡¨
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "æœ¯è¯­æŠ½å–ç»“æœ"
        
        # è®¾ç½®æ ·å¼
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # å†™å…¥æ ‡é¢˜è¡Œï¼ˆåŒè¯­ï¼‰
        headers = ["åºå·", "è‹±æ–‡æœ¯è¯­", "ä¸­æ–‡æœ¯è¯­", "æ¥æºæ–‡ä»¶", "æ¨¡å‹", "Tokenä½¿ç”¨", "å¤„ç†æ—¶é—´"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # å†™å…¥æ•°æ®
        row_num = 2
        term_count = 0
        
        for result in results:
            terms = result.get("extracted_terms", {})
            if "terms" in terms and isinstance(terms["terms"], list):
                for term in terms["terms"]:
                    term_count += 1
                    
                    # å¤„ç†æ¥æºæ–‡ä»¶ä¿¡æ¯
                    source_file = ""
                    if term.get('source_files'):
                        if isinstance(term['source_files'], list):
                            source_file = "; ".join(term['source_files'])
                        else:
                            source_file = str(term['source_files'])
                    elif term.get('source_file'):
                        source_file = term['source_file']
                    elif result.get('source_file'):
                        source_file = result['source_file']
                    
                    # å¤„ç†æ—¶é—´
                    created_time = ""
                    if result.get('created'):
                        from datetime import datetime
                        created_time = datetime.fromtimestamp(result['created']).strftime("%Y-%m-%d %H:%M:%S")
                    
                    # æ”¯æŒåŒè¯­æ ¼å¼
                    eng_term = term.get("eng_term", "")
                    zh_term = term.get("zh_term", "")
                    
                    # å…¼å®¹æ—§æ ¼å¼ï¼ˆå•ä¸€termå­—æ®µï¼‰
                    if not eng_term and not zh_term:
                        single_term = term.get("term", "")
                        # ç®€å•åˆ¤æ–­ï¼šå¦‚æœåŒ…å«ä¸­æ–‡å­—ç¬¦ï¼Œæ”¾å…¥ä¸­æ–‡åˆ—
                        if any('\u4e00' <= char <= '\u9fff' for char in single_term):
                            zh_term = single_term
                        else:
                            eng_term = single_term
                    
                    # å†™å…¥è¡Œæ•°æ®
                    row_data = [
                        term_count,  # åºå·
                        eng_term,  # è‹±æ–‡æœ¯è¯­
                        zh_term,  # ä¸­æ–‡æœ¯è¯­
                        source_file,  # æ¥æºæ–‡ä»¶
                        result.get('model', ''),  # æ¨¡å‹
                        result.get('usage', {}).get('total_tokens', 0),  # Tokenä½¿ç”¨
                        created_time  # å¤„ç†æ—¶é—´
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.border = border
                        # å±…ä¸­å¯¹é½åºå·å’ŒTokenä½¿ç”¨åˆ—
                        if col in [1, 6]:
                            cell.alignment = Alignment(horizontal="center")
                    
                    row_num += 1
            else:
                # å¤„ç†å…¶ä»–æ ¼å¼çš„æ•°æ®
                term_count += 1
                raw_content = terms.get('raw_content', 'æ— å†…å®¹')
                
                row_data = [
                    term_count,
                    raw_content[:100] + "..." if len(raw_content) > 100 else raw_content,
                    result.get('source_file', ''),
                    result.get('model', ''),
                    result.get('usage', {}).get('total_tokens', 0),
                    ""
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.border = border
                    if col in [1, 5]:
                        cell.alignment = Alignment(horizontal="center")
                
                row_num += 1
        
        # è‡ªåŠ¨è°ƒæ•´åˆ—å®½
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            # è®¾ç½®åˆ—å®½ï¼Œæœ€å°10ï¼Œæœ€å¤§50
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # æ·»åŠ ç»Ÿè®¡ä¿¡æ¯å·¥ä½œè¡¨
        stats_ws = wb.create_sheet("ç»Ÿè®¡ä¿¡æ¯")
        stats_data = [
            ["ç»Ÿè®¡é¡¹ç›®", "æ•°å€¼"],
            ["æ€»æœ¯è¯­æ•°", term_count],
            ["å¤„ç†ç»“æœæ•°", len(results)],
            ["æ€»Tokenä½¿ç”¨", sum(r.get('usage', {}).get('total_tokens', 0) for r in results)],
            ["ç”Ÿæˆæ—¶é—´", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = stats_ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # æ ‡é¢˜è¡Œ
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                cell.border = border
        
        # è°ƒæ•´ç»Ÿè®¡ä¿¡æ¯è¡¨åˆ—å®½
        for col in range(1, 3):
            column_letter = get_column_letter(col)
            stats_ws.column_dimensions[column_letter].width = 15
        
        # ä¿å­˜æ–‡ä»¶
        wb.save(output_file)
        
        self.logger.info(f"Excelç»“æœå·²ä¿å­˜åˆ°: {output_file}")
        return str(output_file)
    
    def _save_tbx_results(self, results: List[Dict[str, Any]], filename_base: str) -> str:
        """ä¿å­˜TBXæ ¼å¼ç»“æœ"""
        import xml.etree.ElementTree as ET
        from xml.dom import minidom
        import re
        
        output_file = self.base_dir / f"{filename_base}.tbx"
        
        # æ™ºèƒ½æ£€æµ‹ä¸»è¦è¯­è¨€
        primary_language = self._detect_primary_language(results)
        
        # åˆ›å»ºTBXæ ¹å…ƒç´ 
        root = ET.Element("tbx", attrib={
            "type": "TBX-Default",
            "style": "dct",
            "xml:lang": primary_language,
            "xmlns": "urn:iso:std:iso:30042:ed-2"
        })
        
        # æ·»åŠ TBXå¤´éƒ¨ä¿¡æ¯
        tbx_header = ET.SubElement(root, "tbxHeader")
        
        # æ–‡ä»¶æè¿°
        file_desc = ET.SubElement(tbx_header, "fileDesc")
        title_stmt = ET.SubElement(file_desc, "titleStmt")
        title = ET.SubElement(title_stmt, "title")
        title.text = "å†›äº‹èˆªå¤©æœ¯è¯­åº“"
        
        # å‘å¸ƒä¿¡æ¯
        pub_stmt = ET.SubElement(file_desc, "publicationStmt")
        publisher = ET.SubElement(pub_stmt, "p")
        publisher.text = f"Generated by GPT Term Extraction Tool on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # æºæè¿°
        source_desc = ET.SubElement(file_desc, "sourceDesc")
        source_p = ET.SubElement(source_desc, "p")
        source_p.text = "Extracted from military and aerospace documents using AI processing"
        
        # ç¼–ç æè¿°
        encoding_desc = ET.SubElement(tbx_header, "encodingDesc")
        encoding_p = ET.SubElement(encoding_desc, "p", attrib={"type": "XCSURI"})
        encoding_p.text = "http://www.lisa.org/TBX-Specification.33.0.html"
        
        # ä¿®è®¢æè¿°
        revision_desc = ET.SubElement(tbx_header, "revisionDesc")
        change = ET.SubElement(revision_desc, "change")
        change_date = ET.SubElement(change, "date")
        change_date.text = datetime.now().strftime("%Y-%m-%d")
        change_resp = ET.SubElement(change, "respName")
        change_resp.text = "GPT Term Extraction Tool"
        change_item = ET.SubElement(change, "item", attrib={"type": "create"})
        change_item.text = "Initial term extraction"
        
        # åˆ›å»ºæœ¯è¯­ä½“
        text_body = ET.SubElement(root, "text")
        body = ET.SubElement(text_body, "body")
        
        # å¤„ç†æœ¯è¯­æ•°æ®
        term_count = 0
        for result in results:
            terms = result.get("extracted_terms", {})
            if "terms" in terms and isinstance(terms["terms"], list):
                for term in terms["terms"]:
                    term_count += 1
                    
                    # åˆ›å»ºæœ¯è¯­æ¡ç›®
                    term_entry = ET.SubElement(body, "termEntry", attrib={"id": f"term_{term_count}"})
                    
                    # æ·»åŠ ç®¡ç†ä¿¡æ¯
                    admin_grp = ET.SubElement(term_entry, "adminGrp")
                    admin = ET.SubElement(admin_grp, "admin", attrib={"type": "subjectField"})
                    admin.text = "military_aerospace"
                    
                    # å¤„ç†æ¥æºæ–‡ä»¶ä¿¡æ¯
                    source_file = ""
                    if term.get('source_files'):
                        if isinstance(term['source_files'], list):
                            source_file = "; ".join(term['source_files'])
                        else:
                            source_file = str(term['source_files'])
                    elif term.get('source_file'):
                        source_file = term['source_file']
                    elif result.get('source_file'):
                        source_file = result['source_file']
                    
                    if source_file:
                        source_admin = ET.SubElement(admin_grp, "admin", attrib={"type": "source"})
                        source_admin.text = source_file
                    
                    # æ·»åŠ å¤„ç†ä¿¡æ¯
                    process_admin = ET.SubElement(admin_grp, "admin", attrib={"type": "processStatus"})
                    process_admin.text = "provisionallyProcessed"
                    
                    # æ”¯æŒåŒè¯­æ ¼å¼
                    eng_term = term.get("eng_term", "")
                    zh_term = term.get("zh_term", "")
                    
                    # å…¼å®¹æ—§æ ¼å¼
                    if not eng_term and not zh_term:
                        single_term = term.get('term', '')
                        term_language = self._detect_term_language(single_term)
                        
                        # åˆ›å»ºå•ä¸€è¯­è¨€ç»„
                        lang_grp = ET.SubElement(term_entry, "langGrp", attrib={"xml:lang": term_language})
                        term_grp = ET.SubElement(lang_grp, "termGrp")
                        term_elem = ET.SubElement(term_grp, "term")
                        term_elem.text = single_term
                    else:
                        # åŒè¯­æ ¼å¼ï¼šåˆ›å»ºè‹±æ–‡è¯­è¨€ç»„
                        if eng_term:
                            lang_grp_en = ET.SubElement(term_entry, "langGrp", attrib={"xml:lang": "en"})
                            term_grp_en = ET.SubElement(lang_grp_en, "termGrp")
                            term_elem_en = ET.SubElement(term_grp_en, "term")
                            term_elem_en.text = eng_term
                        
                        # åŒè¯­æ ¼å¼ï¼šåˆ›å»ºä¸­æ–‡è¯­è¨€ç»„
                        if zh_term:
                            lang_grp_zh = ET.SubElement(term_entry, "langGrp", attrib={"xml:lang": "zh"})
                            term_grp_zh = ET.SubElement(lang_grp_zh, "termGrp")
                            term_elem_zh = ET.SubElement(term_grp_zh, "term")
                            term_elem_zh.text = zh_term
                    
                    # æ·»åŠ æå–æ—¶é—´ï¼ˆåœ¨æœ€åä¸€ä¸ªtermGrpä¸­ï¼‰
                    if result.get('created'):
                        # è·å–æœ€ååˆ›å»ºçš„termGrp
                        last_lang_grp = list(term_entry.findall("langGrp"))[-1] if term_entry.findall("langGrp") else None
                        if last_lang_grp is not None:
                            last_term_grp = last_lang_grp.find("termGrp")
                            if last_term_grp is not None:
                                date_admin = ET.SubElement(last_term_grp, "admin", attrib={"type": "created"})
                                date_admin.text = datetime.fromtimestamp(result['created']).strftime("%Y-%m-%d")
        
        # æ ¼å¼åŒ–XMLå¹¶ä¿å­˜
        rough_string = ET.tostring(root, encoding='unicode')
        reparsed = minidom.parseString(rough_string)
        pretty_xml = reparsed.toprettyxml(indent="  ", encoding=None)
        
        # ç§»é™¤ç©ºè¡Œ
        pretty_lines = [line for line in pretty_xml.split('\n') if line.strip()]
        final_xml = '\n'.join(pretty_lines)
        
        # æ·»åŠ XMLå£°æ˜å’ŒDOCTYPE
        xml_declaration = '<?xml version="1.0" encoding="UTF-8"?>\n'
        doctype = '<!DOCTYPE tbx SYSTEM "TBXcoreStructV02.dtd">\n'
        final_content = xml_declaration + doctype + '\n'.join(final_xml.split('\n')[1:])
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(final_content)
        
        self.logger.info(f"TBXç»“æœå·²ä¿å­˜åˆ°: {output_file}")
        return str(output_file)
    
    def _detect_primary_language(self, results: List[Dict[str, Any]]) -> str:
        """æ£€æµ‹æœ¯è¯­çš„ä¸»è¦è¯­è¨€"""
        import re
        
        chinese_count = 0
        english_count = 0
        total_terms = 0
        
        for result in results:
            terms = result.get("extracted_terms", {})
            if "terms" in terms and isinstance(terms["terms"], list):
                for term in terms["terms"]:
                    term_text = term.get('term', '')
                    if term_text:
                        total_terms += 1
                        # ä¼˜å…ˆæ£€æµ‹ä¸­æ–‡ï¼Œå¦‚æœåŒ…å«ä¸­æ–‡å­—ç¬¦å°±ç®—ä¸­æ–‡æœ¯è¯­
                        if re.search(r'[\u4e00-\u9fff]', term_text):  # åŒ…å«ä¸­æ–‡å­—ç¬¦
                            chinese_count += 1
                        # åªæœ‰çº¯è‹±æ–‡æ‰ç®—è‹±æ–‡æœ¯è¯­
                        elif re.search(r'^[a-zA-Z0-9\s\-\(\)\.]+$', term_text):  # çº¯è‹±æ–‡å­—ç¬¦
                            english_count += 1
        
        # æ ¹æ®ä¸»è¦è¯­è¨€è¿”å›è¯­è¨€ä»£ç 
        if chinese_count > english_count:
            return "zh-CN"
        elif english_count > 0:
            return "en"
        else:
            return "zh-CN"  # é»˜è®¤ä¸­æ–‡
    
    def _detect_term_language(self, term_text: str) -> str:
        """æ£€æµ‹å•ä¸ªæœ¯è¯­çš„è¯­è¨€"""
        import re
        
        if not term_text:
            return "zh-CN"
        
        # æ£€æŸ¥æ˜¯å¦åŒ…å«ä¸­æ–‡å­—ç¬¦
        if re.search(r'[\u4e00-\u9fff]', term_text):
            return "zh-CN"
        # æ£€æŸ¥æ˜¯å¦ä¸»è¦æ˜¯è‹±æ–‡å­—ç¬¦
        elif re.search(r'[a-zA-Z]', term_text):
            return "en"
        else:
            return "zh-CN"  # é»˜è®¤ä¸­æ–‡
    
    # =============================================================================
    # æ–­ç‚¹ç»­ä¼ åŠŸèƒ½
    # =============================================================================
    
    def create_processing_checkpoint(self, 
                                   texts: List[str],
                                   source_files: List[str],
                                   processing_config: Dict[str, Any]) -> Optional[str]:
        """
        åˆ›å»ºå¤„ç†æ–­ç‚¹
        
        Args:
            texts: æ–‡æœ¬åˆ—è¡¨
            source_files: æºæ–‡ä»¶åˆ—è¡¨
            processing_config: å¤„ç†é…ç½®
            
        Returns:
            str: æ–­ç‚¹IDï¼Œå¦‚æœæœªå¯ç”¨æ–­ç‚¹åˆ™è¿”å›None
        """
        if not self.enable_checkpoint or not self.checkpoint_manager:
            return None
        
        try:
            # åˆ›å»ºè™šæ‹Ÿæ–‡ä»¶åˆ—è¡¨ï¼ˆåŸºäºæ–‡æœ¬å†…å®¹ï¼‰
            virtual_files = []
            for i, text in enumerate(texts):
                source_file = source_files[i] if i < len(source_files) else f"text_{i+1}.txt"
                virtual_files.append(f"virtual_text_{i}_{source_file}")
            
            checkpoint_id = self.checkpoint_manager.create_checkpoint(
                files=virtual_files,
                processing_config=processing_config,
                output_directory=str(self.base_dir)
            )
            
            self.logger.info(f"åˆ›å»ºå¤„ç†æ–­ç‚¹: {checkpoint_id}")
            return checkpoint_id
            
        except Exception as e:
            self.logger.error(f"åˆ›å»ºæ–­ç‚¹å¤±è´¥: {e}")
            return None
    
    def update_text_processing_status(self, 
                                    text_index: int,
                                    source_file: str,
                                    status: str,
                                    result_data: Optional[Dict] = None,
                                    error_message: Optional[str] = None):
        """
        æ›´æ–°æ–‡æœ¬å¤„ç†çŠ¶æ€
        
        Args:
            text_index: æ–‡æœ¬ç´¢å¼•
            source_file: æºæ–‡ä»¶å
            status: å¤„ç†çŠ¶æ€
            result_data: ç»“æœæ•°æ®
            error_message: é”™è¯¯ä¿¡æ¯
        """
        if not self.enable_checkpoint or not self.checkpoint_manager:
            return
        
        try:
            virtual_file = f"virtual_text_{text_index}_{source_file}"
            self.checkpoint_manager.update_file_status(
                file_path=virtual_file,
                status=status,
                result_data=result_data,
                error_message=error_message
            )
        except Exception as e:
            self.logger.warning(f"æ›´æ–°æ–­ç‚¹çŠ¶æ€å¤±è´¥: {e}")
    
    def list_available_checkpoints(self) -> List[Dict[str, Any]]:
        """åˆ—å‡ºå¯ç”¨çš„æ–­ç‚¹"""
        if not self.enable_checkpoint or not self.checkpoint_manager:
            return []
        
        return self.checkpoint_manager.list_checkpoints()
    
    def load_checkpoint_for_resume(self, checkpoint_id: str) -> bool:
        """åŠ è½½æ–­ç‚¹ç”¨äºç»­ä¼ """
        if not self.enable_checkpoint or not self.checkpoint_manager:
            return False
        
        return self.checkpoint_manager.load_checkpoint(checkpoint_id)
    
    def get_checkpoint_progress(self) -> Optional[Dict[str, Any]]:
        """è·å–å½“å‰æ–­ç‚¹çš„è¿›åº¦ä¿¡æ¯"""
        if not self.enable_checkpoint or not self.checkpoint_manager:
            return None
        
        return self.checkpoint_manager.get_processing_progress()
    
    # =============================================================================
    # å®Œæ•´å¤„ç†æµç¨‹
    # =============================================================================
    
    def run_extraction_only(self, 
                            texts: List[str],
                            system_prompt: str = None,
                            user_prompt_template: str = None,
                            model: str = "gpt-4-turbo-preview",
                            temperature: float = 0.1,
                            max_tokens: int = 4096,
                            max_concurrent: int = 10,
                            description: str = "æœ¯è¯­æŠ½å–æ‰¹å¤„ç†ä»»åŠ¡",
                            source_files: List[str] = None) -> Dict[str, Any]:
        """
        åªè¿è¡ŒæŠ½å–æµç¨‹ï¼Œä¸ä¿å­˜æœ€ç»ˆç»“æœæ–‡ä»¶
        
        Args:
            texts: è¦å¤„ç†çš„æ–‡æœ¬åˆ—è¡¨
            system_prompt: ç³»ç»Ÿæç¤ºè¯
            user_prompt_template: ç”¨æˆ·æç¤ºè¯æ¨¡æ¿
            model: æ¨¡å‹åç§°
            temperature: æ¸©åº¦å‚æ•°
            max_tokens: æœ€å¤§è¾“å‡ºtokenæ•°
            max_concurrent: æœ€å¤§å¹¶å‘æ•°
            description: ä»»åŠ¡æè¿°
            source_files: æ¥æºæ–‡ä»¶ååˆ—è¡¨
            
        Returns:
            Dict[str, Any]: åŒ…å«å¤„ç†ç»“æœçš„å­—å…¸
        """
        self.logger.info("ğŸš€ å¼€å§‹æœ¯è¯­æŠ½å–æµç¨‹")
        
        # åˆ›å»ºå¤„ç†æ–­ç‚¹
        processing_config = {
            "model": model,
            "temperature": temperature,
            "max_tokens": max_tokens,
            "max_concurrent": max_concurrent,
            "description": description
        }
        
        checkpoint_id = self.create_processing_checkpoint(
            texts=texts,
            source_files=source_files or [],
            processing_config=processing_config
        )
        
        try:
            # æ­¥éª¤1: å¹¶å‘å¤„ç†æ‰€æœ‰æ–‡æœ¬
            results = self._run_concurrent_processing(
                texts, system_prompt, user_prompt_template, model, 
                temperature, max_tokens, max_concurrent, source_files
            )
            
            # æ­¥éª¤2: ä¿å­˜åŸå§‹ç»“æœ
            raw_file = self._save_raw_results(results)
            
            # æ­¥éª¤3: æœ¯è¯­å»é‡å¤„ç†
            merged_results = self._run_deduplication(results)
            
            return {
                "raw_results": results,
                "merged_results": merged_results,
                "raw_file": raw_file
            }
            
        except Exception as e:
            self.logger.error(f"âŒ æŠ½å–æµç¨‹æ‰§è¡Œå¤±è´¥: {e}")
            raise

    def run_complete_pipeline(self, 
                             texts: List[str],
                             system_prompt: str = None,
                             user_prompt_template: str = None,
                             model: str = "gpt-4-turbo-preview",
                             temperature: float = 0.1,
                             max_tokens: int = 4096,
                             max_concurrent: int = 10,
                             description: str = "æœ¯è¯­æŠ½å–æ‰¹å¤„ç†ä»»åŠ¡",
                             output_format: str = "json",
                             source_files: List[str] = None) -> Dict[str, str]:
        """
        è¿è¡Œå®Œæ•´çš„æ‰¹å¤„ç†æµç¨‹
        
        Args:
            texts: è¦å¤„ç†çš„æ–‡æœ¬åˆ—è¡¨
            system_prompt: ç³»ç»Ÿæç¤ºè¯
            user_prompt_template: ç”¨æˆ·æç¤ºè¯æ¨¡æ¿
            model: æ¨¡å‹åç§°
            temperature: æ¸©åº¦å‚æ•°
            max_tokens: æœ€å¤§è¾“å‡ºtokenæ•°
            max_concurrent: æœ€å¤§å¹¶å‘æ•°
            description: ä»»åŠ¡æè¿°
            output_format: è¾“å‡ºæ ¼å¼
            source_files: æ¥æºæ–‡ä»¶ååˆ—è¡¨
            
        Returns:
            Dict[str, str]: åŒ…å«å„ä¸ªæ–‡ä»¶è·¯å¾„çš„å­—å…¸
        """
        self.logger.info("ğŸš€ å¼€å§‹å®Œæ•´GPTæ‰¹å¤„ç†æµç¨‹")
        
        # åˆ›å»ºå¤„ç†æ–­ç‚¹
        processing_config = {
            "model": model,
            "temperature": temperature,
            "max_tokens": max_tokens,
            "max_concurrent": max_concurrent,
            "description": description,
            "output_format": output_format
        }
        
        checkpoint_id = self.create_processing_checkpoint(
            texts=texts,
            source_files=source_files or [],
            processing_config=processing_config
        )
        
        try:
            # æ­¥éª¤1: å¹¶å‘å¤„ç†æ‰€æœ‰æ–‡æœ¬
            results = self._run_concurrent_processing(
                texts, system_prompt, user_prompt_template, model, 
                temperature, max_tokens, max_concurrent, source_files
            )
            
            # æ­¥éª¤2: ä¿å­˜åŸå§‹ç»“æœ
            raw_file = self._save_raw_results(results)
            
            # æ­¥éª¤3: æœ¯è¯­å»é‡å¤„ç†
            merged_results = self._run_deduplication(results)
            
            # æ­¥éª¤4: ä¿å­˜æœ€ç»ˆç»“æœ
            # æå–å…ƒæ•°æ®ç”¨äºæ–‡ä»¶å‘½å
            source_filename = self._extract_source_filename(source_files)
            model_name = model.replace("-", "").replace(".", "")  # æ¸…ç†æ¨¡å‹åç”¨äºæ–‡ä»¶å
            total_terms = self._count_total_terms(merged_results)
            
            processed_file = self.save_processed_results(
                merged_results, 
                output_format,
                source_filename,
                model_name,
                total_terms
            )
            
            return {
                "raw_results": raw_file,
                "processed_results": processed_file
            }
            
        except Exception as e:
            self.logger.error(f"âŒ å®Œæ•´æµç¨‹æ‰§è¡Œå¤±è´¥: {e}")
            raise
    
    def _run_concurrent_processing(self, texts: List[str], system_prompt: str, 
                                  user_prompt_template: str, model: str, temperature: float, 
                                  max_tokens: int, max_concurrent: int, 
                                  source_files: List[str]) -> List[Dict[str, Any]]:
        """è¿è¡Œå¹¶å‘å¤„ç†"""
        self.logger.info("ğŸ”„ æ­¥éª¤1: å¹¶å‘å¤„ç†æ–‡æœ¬")
        return self.process_batch_concurrent(
            texts=texts,
            system_prompt=system_prompt,
            user_prompt_template=user_prompt_template,
            model=model,
            temperature=temperature,
            max_tokens=max_tokens,
            max_concurrent=max_concurrent,
            source_files=source_files
        )
    
    def _save_raw_results(self, results: List[Dict[str, Any]]) -> str:
        """ä¿å­˜åŸå§‹ç»“æœ"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        raw_result_file = self.base_dir / f"raw_results_{timestamp}.json"
        with open(raw_result_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        self.logger.info(f"åŸå§‹ç»“æœå·²ä¿å­˜åˆ°: {raw_result_file}")
        return str(raw_result_file)
    
    def _run_deduplication(self, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """è¿è¡Œå»é‡å¤„ç†"""
        self.logger.info("ğŸ”„ æ­¥éª¤2: æœ¯è¯­å»é‡å¤„ç†")
        return self.deduplicate_terms(results)


# =============================================================================
# æ–‡æœ¬åŠ è½½å’Œå¤„ç†å·¥å…·å‡½æ•°
# =============================================================================

def _save_intermediate_text(file_path: str, text: str):
    """ä¿å­˜ä¸­é—´æå–çš„æ–‡æœ¬æ–‡ä»¶"""
    import os
    from pathlib import Path
    
    # åˆ›å»ºextracted_textsæ–‡ä»¶å¤¹
    extracted_dir = Path("extracted_texts")
    extracted_dir.mkdir(exist_ok=True)
    
    # ç”Ÿæˆè¾“å‡ºæ–‡ä»¶å
    source_file = Path(file_path)
    output_file = extracted_dir / f"{source_file.stem}.txt"
    
    try:
        # ä¿å­˜æå–çš„æ–‡æœ¬
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(text)
        
        print(f"ğŸ’¾ ä¸­é—´æ–‡æœ¬å·²ä¿å­˜: {output_file}")
        
    except Exception as e:
        print(f"âš ï¸  ä¿å­˜ä¸­é—´æ–‡æœ¬å¤±è´¥: {e}")

def load_texts_from_file(file_path: str, 
                        chunk_size: Optional[int] = None,
                        use_smart_splitter: bool = True,
                        overlap_size: int = 200) -> List[str]:
    """
    ä»æ–‡ä»¶åŠ è½½æ–‡æœ¬å¹¶è¿›è¡Œæ™ºèƒ½åˆ†å‰²
    
    Args:
        file_path: æ–‡ä»¶è·¯å¾„
        chunk_size: åˆ†å—å¤§å°ï¼ˆå­—ç¬¦æ•°ï¼‰ï¼ŒNoneè¡¨ç¤ºä¸åˆ†å—
        use_smart_splitter: æ˜¯å¦ä½¿ç”¨æ™ºèƒ½åˆ†å‰²å™¨
        overlap_size: é‡å å¤§å°ï¼ˆå­—ç¬¦æ•°ï¼‰
        
    Returns:
        åˆ†å‰²åçš„æ–‡æœ¬åˆ—è¡¨
    """
    from file_processor import FileProcessor
    from text_splitter import TextSplitter
    import os
    
    # åŠ è½½æ–‡ä»¶å†…å®¹
    processor = FileProcessor()
    
    try:
        if file_path.endswith('.pdf'):
            texts = processor.extract_pdf_text(file_path)
        elif file_path.endswith(('.docx', '.doc')):
            texts = processor.extract_docx_text(file_path)
        else:
            # æ–‡æœ¬æ–‡ä»¶
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            texts = [content]
        
        if not texts or not any(text.strip() for text in texts):
            raise ValueError("æ–‡ä»¶å†…å®¹ä¸ºç©ºæˆ–æ— æ³•æå–")
        
        # åˆå¹¶æ‰€æœ‰æ–‡æœ¬
        full_text = '\n\n'.join(texts)
        
        # ä¿å­˜ä¸­é—´æ–‡æœ¬æ–‡ä»¶åˆ°extracted_textsæ–‡ä»¶å¤¹
        _save_intermediate_text(file_path, full_text)
        
        # æ ¹æ®é…ç½®è¿›è¡Œåˆ†å‰²
        if chunk_size and use_smart_splitter:
            # ä½¿ç”¨æ™ºèƒ½åˆ†å‰²å™¨
            max_tokens = max(chunk_size // 4, 500)  # æœ€å°500 tokens
            overlap_tokens = min(overlap_size // 4, max_tokens // 10)  # é‡å ä¸è¶…è¿‡10%
            
            splitter = TextSplitter(
                max_tokens=max_tokens,
                overlap_tokens=overlap_tokens
            )
            return splitter.split_text_with_metadata(full_text, Path(file_path).name)
        elif chunk_size:
            # ç®€å•æŒ‰æ®µè½åˆ†å‰²
            splitter = TextSplitter(max_tokens=10000)  # è®¾ç½®å¾ˆå¤§çš„å€¼ï¼Œé¿å…åˆå¹¶
            chunks = splitter.split_by_paragraphs(full_text)
            
            # æ·»åŠ æ–‡ä»¶æ ‡è¯†
            labeled_chunks = []
            for i, chunk in enumerate(chunks, 1):
                if len(chunks) > 1:
                    labeled_chunk = f"[æ–‡ä»¶: {Path(file_path).name} - ç‰‡æ®µ {i}/{len(chunks)} ({len(chunk)} å­—ç¬¦)]\n{chunk}"
                else:
                    labeled_chunk = f"[æ–‡ä»¶: {Path(file_path).name}]\n{chunk}"
                labeled_chunks.append(labeled_chunk)
            
            return labeled_chunks
        else:
            # ä¸åˆ†å‰²ï¼Œè¿”å›å®Œæ•´æ–‡æœ¬
            return [f"[æ–‡ä»¶: {Path(file_path).name}]\n{full_text}"]
        
    except Exception as e:
        raise ValueError(f"åŠ è½½æ–‡ä»¶å¤±è´¥ {file_path}: {e}")


if __name__ == "__main__":
    # ç®€å•æµ‹è¯•
    print("GPTå¤„ç†å™¨æ¨¡å—åŠ è½½æˆåŠŸ")